from flask import Flask, request, jsonify,send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS
from datetime import datetime, timedelta, date, timezone
import json
import random
import os
import google.generativeai as genai  # Gemini API
from functools import wraps
import io
import PyPDF2  # For reading PDF files
import pandas as pd  # For reading Excel files
from openpyxl import load_workbook  # Alternative for Excel files
from collections import OrderedDict
import base64
from PIL import Image

app = Flask(__name__)

# Enable CORS
CORS(app, resources={r"/*": {"origins": ["http://localhost:3000", "https://samplify.pages.dev"]}})

# MySQL Database Configuration (Using SQLAlchemy)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+mysqlconnector://root:kCqfHlGuizCLUcGqPvcEdtncUWoONxhs@shortline.proxy.rlwy.net:36614/railway'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'ashirwad'  # Change this to a secure secret key

db = SQLAlchemy(app)
jwt = JWTManager(app)
genai.configure(api_key="AIzaSyCyZPcCgke66CyRM_cgBu9EblT6aUbMKqA")


class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%d/%m/%Y %H:%M:%S')
        if isinstance(obj, date):
            return obj.strftime('%d/%m/%Y')
        return super().default(obj)

app.json_encoder = CustomJSONEncoder  # Apply custom encoder


# Add Role model
class Role(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Add AuditLog model
class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)
    action = db.Column(db.String(50), nullable=False)  # e.g., 'CREATE', 'UPDATE', 'DELETE'
    table_name = db.Column(db.String(50), nullable=False)  # e.g., 'Style', 'Task'
    record_id = db.Column(db.Integer, nullable=False)  # ID of the affected record
    old_values = db.Column(db.JSON, nullable=True)  # Previous values
    new_values = db.Column(db.JSON, nullable=True)  # New values
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    ip_address = db.Column(db.String(50), nullable=True)

# ---------------- User Model ----------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(256), nullable=False)
    role_id = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    def has_role(self, role_name):
        return self.role.name == role_name

    def is_admin(self):
        return self.has_role('admin')

# ---------------- Style Model ----------------
class Style(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    style_number = db.Column(db.String(50), nullable=False)
    brand = db.Column(db.String(50), nullable=False)
    sample_type = db.Column(db.String(50), nullable=False)
    garment = db.Column(db.String(50), nullable=False)
    color = db.Column(db.String(50), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    smv = db.Column(db.Float, nullable=True)
    buyer_approval = db.Column(db.Boolean, nullable=True)
    order_received_date = db.Column(db.Date, nullable=True)
    order_delivery_date = db.Column(db.Date, nullable=True)
    approval_status = db.Column(db.Enum("received", "pending", "yetToSend","queued","rejected"), default="queued")
    lab_dips_enabled = db.Column(db.Boolean, nullable=True)
    techpack_data = db.Column(db.JSON, nullable=True)  # JSON column added

    def to_dict(self):
        return {
            "id": self.id,
            "style": self.style_number,
            "buyer": self.brand,
            "garment": self.garment,
            "date": self.order_received_date    ,
            "approvalStatus": self.approval_status,
        }


# ---------------- Activity Model ----------------
class Activity(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    style = db.Column(db.String(50), nullable=False)
    process = db.Column(db.String(100), nullable=False)
    duration = db.Column(db.Float, nullable=False)
    planned_start = db.Column(db.String(20), nullable=True)
    planned_end = db.Column(db.String(20), nullable=True)
    actual_start = db.Column(db.String(20), nullable=True)
    actual_end = db.Column(db.String(20), nullable=True)
    actual_duration = db.Column(db.Float, nullable=True)
    delay = db.Column(db.Float, nullable=True)
    responsibility = db.Column(db.String(50), nullable=False)
class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    style_number = db.Column(db.String(50), nullable=False)
    brand = db.Column(db.String(50), nullable=False)
    sample_type = db.Column(db.String(50), nullable=False)
    garment = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(20), default="toBeDone")
    progress = db.Column(db.Integer, default=0)
    priority = db.Column(db.String(20), default="Medium")
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    comment = db.Column(db.Text, nullable=True)
    problem_reported = db.Column(db.Boolean, default=False)

    steps = db.relationship("TaskStep", backref="task", cascade="all, delete", lazy=True,order_by="TaskStep.id")

# Task Step Model 
class TaskStep(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    task_id = db.Column(db.Integer, db.ForeignKey("task.id"), nullable=False)
    step_name = db.Column(db.String(50), nullable=False)
    is_completed = db.Column(db.Boolean, default=False)

class Courier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    style_number = db.Column(db.String(50), nullable=False)
    courier_name = db.Column(db.String(100), nullable=False)
    awb_number = db.Column(db.String(50), nullable=False)
    att = db.Column(db.String(50), nullable=False)
    content = db.Column(db.String(255), nullable=False)
    garment_type = db.Column(db.String(50), nullable=False)
    placement_date = db.Column(db.DateTime, default=datetime.utcnow)
    completion_date = db.Column(db.DateTime, nullable=True)

# Order Status Model
class OrderStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    courier_id = db.Column(db.Integer, db.ForeignKey('courier.id'), nullable=False)
    type = db.Column(db.String(100), nullable=False)
    completed = db.Column(db.Boolean, default=False)
    placement_date = db.Column(db.DateTime, default=datetime.utcnow)
    completion_date = db.Column(db.DateTime, nullable=True)

class LabDip(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    style_number = db.Column(db.String(50), nullable=False)
    buyer = db.Column(db.String(50))
    fabric = db.Column(db.String(50))
    color = db.Column(db.String(50))
    shade = db.Column(db.String(10), default='A')
    status = db.Column(db.Enum('approved', 'pending', 'yetToSend'), default='yetToSend')
    approval_date = db.Column(db.Date)

    def to_dict(self):
        return {
            "id": self.id,
            "style": self.style_number,
            "buyer": self.buyer,
            "fabric": self.fabric,
            "color": self.color,
            "shade": self.shade,
            "status": self.status,
            "date": self.approval_date.strftime('%d/%m/%Y') if self.approval_date else None
        }

class Trim(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), unique=True, nullable=False)
    image = db.Column(db.String(255), nullable=True)
    composition = db.Column(db.String(255), nullable=True)
    structure = db.Column(db.String(255), nullable=True)
    shade = db.Column(db.String(255), nullable=True)
    brand = db.Column(db.String(255), nullable=True)
    code = db.Column(db.String(255), nullable=True)

class TrimVariant(db.Model):
    __tablename__ = 'trim_variants'
    id = db.Column(db.Integer, primary_key=True)
    trim_id = db.Column(db.Integer)
    image = db.Column(db.String(255), nullable=False)
    composition = db.Column(db.String(255))
    structure = db.Column(db.String(255))
    shade = db.Column(db.String(255))
    brand = db.Column(db.String(255))
    code = db.Column(db.String(255))
    rate = db.Column(db.String(255))


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    message = db.Column(db.String(255), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    read_status = db.Column(db.Boolean, default=False)

    def to_dict(self):
        return {
            "id": self.id,
            "message": self.message,
            "timestamp": self.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "read_status": self.read_status
        }

class SampleTrackerStyle(db.Model):
    __tablename__ = 'sample_tracker_style'
    id = db.Column(db.Integer, primary_key=True)
    style_number = db.Column(db.String(100), nullable=False)
    brand = db.Column(db.String(100), nullable=False)
    garment_type = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    end_date = db.Column(db.DateTime, nullable=True)
    # Relationship: A style has many samples. Cascade delete means if a style is deleted, its samples are too.
    samples = db.relationship('SampleTrackerSample', backref='style', lazy=True, cascade="all, delete-orphan")

    def to_dict(self, include_samples=True):
        data = {
            "id": self.id,
            "styleNumber": self.style_number,
            "brand": self.brand,
            "garmentType": self.garment_type,
            "startDate": self.start_date.isoformat() if self.start_date else None,
            "endDate": self.end_date.isoformat() if self.end_date else None,
        }
        if include_samples:
            # Sort samples by their ID to maintain a somewhat consistent order, or by another field if needed
            data["samples"] = sorted([sample.to_dict() for sample in self.samples], key=lambda s: s['id'])
        return data

class SampleTrackerSample(db.Model):
    __tablename__ = 'sample_tracker_sample'
    id = db.Column(db.Integer, primary_key=True) # Unique ID for each sample
    style_id = db.Column(db.Integer, db.ForeignKey('sample_tracker_style.id'), nullable=False)
    type = db.Column(db.String(100), nullable=False)
    completed = db.Column(db.Boolean, default=False)
    start_date = db.Column(db.DateTime, nullable=True)
    end_date = db.Column(db.DateTime, nullable=True)

    def to_dict(self):
        return {
            "id": self.id,
            "styleId": self.style_id, # Optional: useful for context
            "type": self.type,
            "completed": self.completed,
            "startDate": self.start_date.isoformat() if self.start_date else None,
            "endDate": self.end_date.isoformat() if self.end_date else None
        }

# New Fabric models
class Fabric(db.Model):
    __tablename__ = 'fabrics'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    image = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship with variants
    variants = db.relationship('FabricVariant', backref='fabric', lazy=True, cascade='all, delete-orphan')

class FabricVariant(db.Model):
    __tablename__ = 'fabric_variants'
    
    id = db.Column(db.Integer, primary_key=True)
    fabric_id = db.Column(db.Integer, db.ForeignKey('fabrics.id'), nullable=False)
    image = db.Column(db.String(255), nullable=True)
    composition = db.Column(db.String(255), nullable=True)
    structure = db.Column(db.String(255), nullable=True)
    shade = db.Column(db.String(100), nullable=True)
    brand = db.Column(db.String(100), nullable=True)
    code = db.Column(db.String(100), nullable=True)
    rate = db.Column(db.String(100), nullable=True)
    supplier = db.Column(db.String(100), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

PREDEFINED_SAMPLE_TYPES = ['Fit Sample', 'PP Sample', 'SMS', 'Photoshoot Sample', 'TOP Sample', 'FOB Sample']


PREDEFINED_PROCESSES = [
    {"process": "Order Receipt (Buyer PO)", "duration": 0, "responsibility": "Merchandiser"},
    {"process": "CAD Consumption Received", "duration": 2, "responsibility": "CAD Department"},
    {"process": "BOM Generation", "duration": 0.5, "responsibility": "Merchandiser"},
    {"process": "PO Issue for Fabric & Trims", "duration": 0.5, "responsibility": "Merchandiser"},
    {"process": "Fabric Received", "duration": 34, "responsibility": "Store Manager"},
    {"process": "Sample Indent Made", "duration": 0.5, "responsibility": "Merchandiser"},
    {"process": "Pattern Cutting", "duration": 0.5, "responsibility": "Pattern Master"},
    {"process": "Sewing", "duration": 4, "responsibility": "Production Head"},
    {"process": "Embroidery", "duration": 0.5, "responsibility": "Embroidery Head"},
    {"process": "Finishing", "duration": 0.5, "responsibility": "Production Head"},
    {"process": "Packing", "duration": 1, "responsibility": "Production Head"},
    {"process": "Documentation in PLM", "duration": 0.5, "responsibility": "Production Head"},
    {"process": "Dispatch", "duration": 0.5, "responsibility": "Merchandiser"},
]

# Create tables in the database
with app.app_context():
    db.create_all()

# ---------------- Authentication Routes ----------------
@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role_name = data.get('role', 'user')  # Default to 'user' role if not specified

    if User.query.filter_by(username=username).first():
        return jsonify({"message": "User already exists"}), 400

    role = Role.query.filter_by(name=role_name).first()
    if not role:
        return jsonify({"message": "Invalid role"}), 400

    hashed_password = generate_password_hash(password)
    new_user = User(
        username=username,
        password=hashed_password,
        role_id=role.id
    )
    db.session.add(new_user)
    db.session.commit()

    # Log the user creation
    log_audit(
        user_id=new_user.id,
        action='CREATE',
        table_name='User',
        record_id=new_user.id,
        new_values={'username': username, 'role': role_name},
        ip_address=request.remote_addr
    )

    return jsonify({"message": "User registered successfully"}), 201

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    user = User.query.filter_by(username=username).first()

    if not user or not check_password_hash(user.password, password):
        return jsonify({"message": "Invalid credentials"}), 401

    if not user.is_active:
        return jsonify({"message": "Account is deactivated"}), 403

    # Update last login
    user.last_login = datetime.utcnow()
    db.session.commit()

    access_token = create_access_token(identity=user.id)
    return jsonify({
        "access_token": access_token,
        "role": user.role_id,
        "username": user.username
    }), 200

# ---------------- Styles Routes (Using SQLAlchemy) ----------------
@app.route('/styles', methods=['GET'])
def get_styles():
    styles = Style.query.all()
    return jsonify([{
        "id": s.id,
        "styleNumber": s.style_number,
        "brand": s.brand,
        "sampleType": s.sample_type,
        "garment": s.garment,
        "color": s.color,
        "quantity": s.quantity,
        "smv": s.smv,
        "buyerApproval": s.buyer_approval,
        "orderReceivedDate": s.order_received_date.strftime("%Y-%m-%d") if s.order_received_date else None,
        "orderDeliveryDate": s.order_delivery_date.strftime("%Y-%m-%d") if s.order_delivery_date else None,
        "labDipsEnabled" : s.lab_dips_enabled,
        "techpackData" : s.techpack_data
    } for s in styles])

@app.route('/styles', methods=['POST'])
def add_style(): 
    data = request.json
    new_style = Style(
        style_number=data['styleNumber'],
        brand=data['brand'],
        sample_type=data['sampleType'],
        garment=data['garment'],
        color=data['color'],
        quantity=data['quantity'],
        smv=data.get('smv'),
        buyer_approval=data.get('buyerApproval'),
        order_received_date=datetime.strptime(data['orderReceivedDate'], "%Y-%m-%d") if data.get('orderReceivedDate') else None,
        order_delivery_date=datetime.strptime(data['orderDeliveryDate'], "%Y-%m-%d") if data.get('orderDeliveryDate') else None,
        lab_dips_enabled = data.get("labDipsEnabled")
    )
    db.session.add(new_style)
    db.session.commit()
    new_task = Task(
        style_number=data["styleNumber"],
        brand=data["brand"],
        sample_type=data["sampleType"],
        garment=data["garment"],
        status=data.get("status", "toBeDone"),
    )
    db.session.add(new_task)
    db.session.commit()
    # Add steps for the task
    steps = ["Start","Indent","PatternCutting","FabricCutting","Sewing","Embroidery","Finishing", "Problem"]
    process_names = [item["process"] for item in PREDEFINED_PROCESSES]
    process_names.insert(0,"Start")
    process_names.append("Problem")
    # steps = ["start", "indent", "pattern", "fabric", "embroidery", "packing", "problem"]
    for step in process_names:
        db.session.add(TaskStep(task_id=new_task.id, step_name=step, is_completed=False))
    db.session.commit()
    if data.get("labDipsEnabled") == True:
        new_lab_dip = LabDip(
        style_number=data['styleNumber'],
        buyer=data['brand'],
        fabric=data.get('fabric', 'Cotton Blend'),
        color=data['color'],
        shade=data.get('shade', 'A'),
        status=data.get('status', 'yetToSend'),
        approval_date=datetime.strptime(data['date'], '%d/%m/%Y') if 'date' in data else None
    )
        db.session.add(new_lab_dip)
        db.session.commit()


    return jsonify({"message": "Style added successfully", "id": new_style.id}), 201

@app.route('/styles/<int:style_id>', methods=['DELETE'])
def delete_style(style_id):
    style = Style.query.get(style_id)
    if style:
        db.session.delete(style)
        db.session.commit()
        return jsonify({'message': 'Style deleted successfully'}), 200
    return jsonify({'error': 'Style not found'}), 404


@app.route('/styles/<int:style_id>', methods=['PUT'])
def update_style(style_id):
    style = Style.query.get(style_id)
    if not style:
        return jsonify({"error": "Style not found"}), 404
    data = request.json
    style.style_number = data.get('styleNumber', style.style_number)
    style.brand = data.get('brand', style.brand)
    style.sample_type = data.get('sampleType', style.sample_type)
    style.garment = data.get('garment', style.garment)
    style.color = data.get('color', style.color)
    style.quantity = data.get('quantity', style.quantity)
    style.smv = data.get('smv', style.smv)
    style.buyer_approval = data.get('buyerApproval', style.buyer_approval)
    style.order_received_date = datetime.strptime(data['orderReceivedDate'], "%Y-%m-%d") if data.get('orderReceivedDate') else style.order_received_date
    style.order_delivery_date = datetime.strptime(data['orderDeliveryDate'], "%Y-%m-%d") if data.get('orderDeliveryDate') else style.order_delivery_date
    style.lab_dips_enabled = data.get("labDipsEnabled", style.lab_dips_enabled)
    if(style.buyer_approval):
        style.approval_status = "received"
    db.session.commit()

    task = Task.query.filter(Task.style_number == style.style_number).first()
    if task is None:
        new_task = Task(
        style_number=data["styleNumber"],
        brand=data["brand"],
        sample_type=data["sampleType"],
        garment=data["garment"],
        status=data.get("status", "toBeDone"),
        )
        db.session.add(new_task)
        db.session.commit()
        # Add steps for the task
        steps = ["Start","Indent","PatternCutting","FabricCutting","Sewing","Embroidery","Finishing", "Problem"]
        # steps = ["start", "indent", "pattern", "fabric", "embroidery", "packing", "problem"]
        process_names = [item["process"] for item in PREDEFINED_PROCESSES]
        process_names.insert(0,"Start")
        process_names.append("Problem")
        for step in process_names:
            db.session.add(TaskStep(task_id=new_task.id, step_name=step, is_completed=False))
        db.session.commit()


    if data.get("labDipsEnabled") == True:
        new_lab_dip = LabDip(
        style_number=style.style_number,
        buyer=style.brand,
        fabric=style.garment,
        color=style.color,
        shade=data.get('shade', 'A'),
        status=data.get('status', 'yetToSend'),
        approval_date=datetime.strptime(style.order_received_date, '%d/%m/%Y') if 'date' in data else None
        )
        db.session.add(new_lab_dip)
        db.session.commit()

    sample_tracking_sample = SampleTrackerStyle.query.filter_by(style_number=data['styleNumber']).first()

    if not sample_tracking_sample:
        new_style = SampleTrackerStyle(
            style_number=data['styleNumber'],
            brand=data['brand'],
            garment_type=data['garment'],
            start_date=datetime.now(timezone.utc) # Set start date on creation
        )
        db.session.add(new_style)
        db.session.flush()  # Flush to get the new_style.id for linking samples

        # Add predefined samples to this new style
        for sample_type_name in PREDEFINED_SAMPLE_TYPES:
            sample = SampleTrackerSample(
                style_id=new_style.id,
                type=sample_type_name,
                completed=False
                # start_date and end_date will be null initially
            )
            db.session.add(sample)

        db.session.commit()

    return jsonify({"message": "Style updated successfully"}), 200
#----------------- Sample Tracking ---------------------


# --- API Endpoints for Sample Tracker ---

@app.route('/api/sample-tracker/sample-types', methods=['GET'])
def get_predefined_sample_types():
    """Returns the list of predefined sample types."""
    return jsonify(PREDEFINED_SAMPLE_TYPES)

@app.route('/api/sample-tracker/styles', methods=['GET'])
def get_all_sample_tracker_styles():
    """Gets all sample tracker styles and their associated samples."""
    styles = SampleTrackerStyle.query.order_by(SampleTrackerStyle.start_date.desc()).all()
    return jsonify([s.to_dict() for s in styles])

@app.route('/api/sample-tracker/styles', methods=['POST'])
def create_sample_tracker_style():
    """Creates a new style and initializes it with predefined samples."""
    data = request.json
    if not data or not data.get('styleNumber') or not data.get('brand') or not data.get('garmentType'):
        return jsonify({"error": "Missing required fields: styleNumber, brand, garmentType"}), 400

    new_style = SampleTrackerStyle(
        style_number=data['styleNumber'],
        brand=data['brand'],
        garment_type=data['garmentType'],
        start_date=datetime.now(timezone.utc) # Set start date on creation
    )
    db.session.add(new_style)
    db.session.flush()  # Flush to get the new_style.id for linking samples

    # Add predefined samples to this new style
    for sample_type_name in PREDEFINED_SAMPLE_TYPES:
        sample = SampleTrackerSample(
            style_id=new_style.id,
            type=sample_type_name,
            completed=False
            # start_date and end_date will be null initially
        )
        db.session.add(sample)

    db.session.commit()
    return jsonify(new_style.to_dict()), 201

@app.route('/api/sample-tracker/styles/<int:style_id>', methods=['DELETE'])
def remove_sample_tracker_style(style_id):
    """Deletes a style and its associated samples."""
    style = SampleTrackerStyle.query.get(style_id)
    if not style:
        return jsonify({"error": "Style not found"}), 404

    db.session.delete(style)  # Cascade delete will handle samples
    db.session.commit()
    return jsonify({"message": "Style and its samples deleted successfully"}), 200

@app.route('/api/sample-tracker/styles/<int:style_id>/samples', methods=['POST'])
def add_custom_sample_to_style(style_id):
    """Adds a new custom sample to an existing style."""
    style = SampleTrackerStyle.query.get(style_id)
    if not style:
        return jsonify({"error": "Style not found"}), 404

    data = request.json
    sample_type = data.get('type')
    if not sample_type or not sample_type.strip():
        return jsonify({"error": "Sample type is required and cannot be empty"}), 400

    new_sample = SampleTrackerSample(
        style_id=style.id,
        type=sample_type.strip(),
        completed=False
    )
    db.session.add(new_sample)
    db.session.commit()
    # Return the updated style object
    return jsonify(style.to_dict()), 201


@app.route('/api/sample-tracker/styles/<int:style_id>/samples/<int:sample_id>', methods=['DELETE'])
def remove_sample_from_style(style_id, sample_id):
    """Deletes a specific sample from a style."""
    sample = SampleTrackerSample.query.filter_by(id=sample_id, style_id=style_id).first()
    if not sample:
        return jsonify({"error": "Sample not found or does not belong to the specified style"}), 404

    db.session.delete(sample)
    db.session.commit()
    # Return the updated parent style object
    style = SampleTrackerStyle.query.get(style_id)
    return jsonify(style.to_dict()), 200


@app.route('/api/sample-tracker/styles/<int:style_id>/samples/<int:sample_id>/toggle', methods=['PUT'])
def update_sample_status(style_id, sample_id):
    """Toggles the completion status of a sample and updates its dates."""
    sample = SampleTrackerSample.query.filter_by(id=sample_id, style_id=style_id).first()
    if not sample:
        return jsonify({"error": "Sample not found or does not belong to the specified style"}), 404

    sample.completed = not sample.completed
    current_time = datetime.now(timezone.utc)

    if sample.completed:
        if not sample.start_date: # Set start_date only if it's not already set
            sample.start_date = current_time
        sample.end_date = current_time
    else:
        # When un-completing, nullify the end_date.
        # Decide if start_date should also be nullified or preserved. Preserving is common.
        sample.end_date = None
        # If you want to reset start_date as well when un-completing:
        # sample.start_date = None

    db.session.commit() # Commit sample changes first

    # After updating the sample, check if the parent style is now fully completed
    style = SampleTrackerStyle.query.get(style_id)
    if style:
        all_samples_completed = all(s.completed for s in style.samples)
        if all_samples_completed:
            if not style.end_date: # Set style end_date only if not already set
                 style.end_date = current_time # Or max(s.end_date for s in style.samples if s.end_date)
        else:
            style.end_date = None # If any sample is not complete, style is not complete
        db.session.commit()

    # Return the updated parent style object
    return jsonify(style.to_dict()), 200



# ---------------- Activity Routes (Using SQLAlchemy) ----------------
@app.route("/api/activity", methods=["GET"])
def get_activities():
    style = request.args.get("style")
    if not style:
        return jsonify({"error": "Style parameter is required"}), 400

    activities = Activity.query.filter_by(style=style).all()
    if not activities:
        return jsonify({"message": "No activities found for the given style"}), 404

    return jsonify([
        {
            "id": act.id,
            "style": act.style,
            "process": act.process,
            "duration": act.duration,
            "plannedStart": act.planned_start,
            "plannedEnd": act.planned_end,
            "actualStart": act.actual_start,
            "actualEnd": act.actual_end,
            "actualDuration": act.actual_duration,
            "delay": act.delay,
            "responsibility": act.responsibility
        } 
        for act in activities
    ])

@app.route("/api/activity", methods=["POST"])
def add_activities():
    data = request.json
    style = data.get("style")
    received_date = data.get("receivedDate")

    if not style:
        return jsonify({"error": "Style is required"}), 400
    if not received_date:
        return jsonify({"error": "Received Date is required"}), 400

    # Check if activities for this style already exist
    if Activity.query.filter_by(style=style).first():
        return jsonify({"message": "Activities for this style already exist"}), 400

    try:
        planned_start = datetime.strptime(received_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

    for process in PREDEFINED_PROCESSES:
        planned_end = planned_start + timedelta(days=process["duration"])

        new_activity = Activity(
            style=style,
            process=process["process"],
            duration=process["duration"],
            planned_start=planned_start.strftime("%Y-%m-%d"),
            planned_end=planned_end.strftime("%Y-%m-%d"),
            actual_start=datetime.strptime(received_date, "%Y-%m-%d") if process["process"] == "Order Receipt (Buyer PO)" else None,
            actual_end=None,
            actual_duration=None,
            delay=None,
            responsibility=process["responsibility"]
        )

        db.session.add(new_activity)
        planned_start = planned_end  # Update for next process

    db.session.commit()
    return jsonify({"message": "Activities initialized for style"}), 201

@app.route("/api/activity/<int:id>", methods=["PUT"])
def update_activity(id):
    activity = Activity.query.get_or_404(id)

    data = request.json
    actual_start = data.get("actualStart")
    actual_end = data.get("actualEnd")

    if actual_start:
        try:
            # First try parsing with just the date
            actual_start_date = datetime.strptime(actual_start, "%Y-%m-%d")
            activity.actual_start = actual_start_date
        except ValueError:
            try:
                # If that fails, try with time component
                actual_start_date = datetime.strptime(actual_start, "%Y-%m-%d %H:%M:%S")
                activity.actual_start = actual_start
            except ValueError:
                return jsonify({"error": "Invalid actual_start format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS"}), 400

    if actual_end:
        try:
            # First try parsing with just the date
            actual_end_date = datetime.strptime(actual_end, "%Y-%m-%d")
            activity.actual_end = actual_end
        except ValueError:
            try:
                # If that fails, try with time component
                actual_end_date = datetime.strptime(actual_end, "%Y-%m-%d %H:%M:%S")
                activity.actual_end = actual_end
            except ValueError:
                return jsonify({"error": "Invalid actual_end format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS"}), 400

    if activity.actual_start and activity.actual_end:
        # Parse dates with flexible formats
        try:
            # Try parsing with date format first
            start_date = datetime.strptime(activity.actual_start, "%Y-%m-%d")
        except ValueError:
            try:
                # Try with time component if date-only format fails
                start_date = datetime.strptime(activity.actual_start, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return jsonify({"error": f"Could not parse actual_start date: {activity.actual_start}"}), 400
                
        try:
            # Try parsing with date format first
            end_date = datetime.strptime(activity.actual_end, "%Y-%m-%d")
        except ValueError:
            try:
                # Try with time component if date-only format fails
                end_date = datetime.strptime(activity.actual_end, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return jsonify({"error": f"Could not parse actual_end date: {activity.actual_end}"}), 400
        
        # Calculate duration
        activity.actual_duration = (end_date - start_date).days

        if activity.planned_end:
            try:
                # Try to parse the planned_end date with date-only format
                planned_end_date = datetime.strptime(activity.planned_end, "%Y-%m-%d")
                activity.delay = max((end_date - planned_end_date).days, 0)
            except ValueError:
                try:
                    # Try with time component if date-only format fails
                    planned_end_date = datetime.strptime(activity.planned_end, "%Y-%m-%d %H:%M:%S")
                    activity.delay = max((end_date - planned_end_date).days, 0)
                except ValueError:
                    print(f"Could not parse planned_end date: {activity.planned_end}")
                    activity.delay = None

    db.session.commit()
    return jsonify({"message": "Activity updated successfully"}), 200

@app.route("/api/approval-status", methods=["GET"])
def get_approval_status():
    styles = Style.query.all()
    
    categorized_samples = {"received": [], "pending": [], "yetToSend": [], "queued" : [], "rejected" : []}
    
    for style in styles:
        categorized_samples[style.approval_status].append(style.to_dict())

    return jsonify(categorized_samples)

# API to update approval status of a style
@app.route("/api/update-status", methods=["POST"])
def update_status():
    data = request.json
    style_id = data.get("id")
    new_status = data.get("approvalStatus")

    if new_status not in ["received", "pending", "yetToSend", "rejected"]:
        return jsonify({"error": "Invalid status"}), 400

    style = Style.query.get(style_id)
    if not style:
        return jsonify({"error": "Style not found"}), 404

    style.approval_status = new_status
    db.session.commit()

    return jsonify({"message": "Status updated successfully", "style": style.to_dict()})

@app.route("/tasks", methods=["POST"])
def add_task():
    data = request.json
    new_task = Task(
        style_number=data["styleNumber"],
        brand=data["brand"],
        sample_type=data["sampleType"],
        garment=data["garment"],
        status=data.get("status", "toBeDone"),
    )
    db.session.add(new_task)
    db.session.commit()

    # Add steps for the task
    # steps = ["start", "indent", "pattern", "fabric", "embroidery", "packing", "problem"]
    steps = ["Start","Indent","PatternCutting","FabricCutting","Sewing","Embroidery","Finishing", "Problem"]
    process_names = [item["process"] for item in PREDEFINED_PROCESSES]
    process_names.insert(0,"Start")
    process_names.append("Problem")
    for step in process_names:
        db.session.add(TaskStep(task_id=new_task.id, step_name=step, is_completed=False))
    db.session.commit()

    return jsonify({"message": "Task added successfully", "task_id": new_task.id}), 201

# Get all tasks
@app.route("/tasks", methods=["GET"])
def get_tasks():
    tasks = Task.query.all()
    task_list = []
    for task in tasks:
        ordered_steps = sorted(task.steps, key=lambda s: s.id)
        task_list.append({
            "id": task.id,
            "styleNumber": task.style_number,
            "brand": task.brand,
            "sampleType": task.sample_type,
            "garment": task.garment,
            "status": task.status,
            "progress": task.progress,
            "priority": task.priority,
            "timestamp": task.timestamp,
            "comment": task.comment,
            "problemReported": task.problem_reported,
            "steps":  [
                {"step_name": step.step_name, "is_completed": step.is_completed}
                for step in ordered_steps
            ]
        })
    app.logger.info(task_list)
    return jsonify(task_list)

# Get a single task
@app.route("/tasks/<int:task_id>", methods=["GET"])
def get_task(task_id):
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    ordered_steps = sorted(task.steps, key=lambda s: s.id)
    task_data = {
        "id": task.id,
        "styleNumber": task.style_number,
        "brand": task.brand,
        "sampleType": task.sample_type,
        "garment": task.garment,
        "status": task.status,
        "progress": task.progress,
        "priority": task.priority,
        "timestamp": task.timestamp,
        "comment": task.comment,
        "problemReported": task.problem_reported,
        "steps": {step.step_name: step.is_completed for step in ordered_steps},
    }
    return jsonify(task_data)

# Update a task (progress, status, steps, comments)
@app.route("/tasks/<int:task_id>", methods=["PUT"])
def update_task(task_id):
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    data = request.json
    if "status" in data:
        task.status = data["status"]
    if "progress" in data:
        task.progress = data["progress"]
    if "comment" in data:
        task.comment = data["comment"]
    if "problemReported" in data:
        task.problem_reported = data["problemReported"]

    # Update steps
    if "steps" in data:
        for item  in data["steps"]:
            step = TaskStep.query.filter_by(task_id=task_id, step_name=item["step_name"]).first()
            if step:
                step.is_completed = item["is_completed"]

    db.session.commit()
    if "status" in data and data["status"] == "completed":
        courier = Courier(
        style_number=task.style_number,
        courier_name="Blue Dart",
        awb_number="#2025" + str(random.randint(1,20000)),
        att="#ASJK" + str(random.randint(1,50)),
        content=task.sample_type,
        garment_type=task.garment
    )
        db.session.add(courier)
        db.session.commit()

        for order_type in ['Booked Portal', 'Sent from Factory', 'In Transit', 'Received']:
            order_status = OrderStatus(
                courier_id=courier.id,
                type=order_type
            )
            db.session.add(order_status)
        
        db.session.commit()
        style = Style.query.filter_by(style_number=task.style_number).first()
        if style is not None:
            style.approval_status = "yetToSend"
            db.session.commit()
        

    return jsonify({"message": "Task updated successfully"})

# Delete a task
@app.route("/tasks/<int:task_id>", methods=["DELETE"])
def delete_task(task_id):
    task = Task.query.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    db.session.delete(task)
    db.session.commit()
    return jsonify({"message": "Task deleted successfully"})

@app.route('/add_courier', methods=['POST'])
def add_courier():
    data = request.json
    courier = Courier(
        style_number=data['styleNumber'],
        courier_name=data['courierName'],
        awb_number=data['awbNumber'],
        att=data['att'],
        content=data['content'],
        garment_type=data['garmentType']
    )
    db.session.add(courier)
    db.session.commit()

    for order_type in ['Booked Portal', 'Sent from Factory', 'In Transit', 'Received']:
        order_status = OrderStatus(
            courier_id=courier.id,
            type=order_type
        )
        db.session.add(order_status)
    
    db.session.commit()
    return jsonify({'message': 'Courier added successfully'}), 201

@app.route('/get_couriers', methods=['GET'])
def get_couriers():
    couriers = Courier.query.all()
    result = []
    for courier in couriers:
        orders = OrderStatus.query.filter_by(courier_id=courier.id).all()
        order_list = [
            {
                "id": order.id,
                "type": order.type,
                "completed": order.completed,
                "placement": order.placement_date.strftime('%d/%m/%Y'),
                "date": order.completion_date
            }
            for order in orders
        ]
        result.append({
            "id": courier.id,
            "styleNumber": courier.style_number,
            "courierName": courier.courier_name,
            "awbNumber": courier.awb_number,
            "att": courier.att,
            "content": courier.content,
            "garmentType": courier.garment_type,
            "placement": courier.placement_date,
            "date": courier.completion_date,
            "orders": order_list
        })
    return jsonify(result), 200

@app.route('/update_order', methods=['POST'])
def update_order():
    data = request.json
    order_id = data.get('orderId')  # Extract orderId
    completed = data.get('completed')

    if order_id is None or completed is None:
        return jsonify({'error': 'Missing required fields'}), 400

    order = OrderStatus.query.get(order_id)

    if not order:
        return jsonify({'error': 'Order not found'}), 404

    # Update order status
    order.completed = completed
    order.completion_date = datetime.utcnow() if completed else None

    db.session.commit()
    if order.type=="Sent from Factory":
        courier = Courier.query.filter_by(id=order.courier_id).first()
        style = Style.query.filter_by(style_number=courier.style_number).first()
        if style:
            style.approval_status = "pending"
            db.session.commit()

    return jsonify({
        'message': 'Order updated successfully',
        'orderId': order.id,
        'completed': order.completed,
        'completionDate': order.completion_date.strftime('%d/%m/%Y') if order.completion_date else None
    }), 200

@app.route('/delete_courier/<int:courier_id>', methods=['DELETE'])
def delete_courier(courier_id):
    courier = Courier.query.get(courier_id)
    if courier:
        OrderStatus.query.filter_by(courier_id=courier_id).delete()
        db.session.delete(courier)
        db.session.commit()
        return jsonify({'message': 'Courier deleted successfully'}), 200
    return jsonify({'error': 'Courier not found'}), 404

@app.route('/lab_dips', methods=['GET'])
def get_lab_dips():
    lab_dips = LabDip.query.all()
    response = {
        "approved": [dip.to_dict() for dip in lab_dips if dip.status == "approved"],
        "pending": [dip.to_dict() for dip in lab_dips if dip.status == "pending"],
        "yetToSend": [dip.to_dict() for dip in lab_dips if dip.status == "yetToSend"]
    }
    return jsonify(response), 200

@app.route('/lab_dips', methods=['POST'])
def add_lab_dip():
    data = request.json
    style = Style.query.filter_by(style_no=data['style']).first()
    
    if not style:
        style = Style(style_no=data['style'])
        db.session.add(style)
        db.session.commit()

    new_lab_dip = LabDip(
        style_id=style.id,
        buyer=data['buyer'],
        fabric=data['fabric'],
        color=data['color'],
        shade=data.get('shade', 'A'),
        status=data.get('status', 'yetToSend'),
        approval_date=datetime.strptime(data['date'], '%d/%m/%Y') if 'date' in data else None
    )
    db.session.add(new_lab_dip)
    db.session.commit()

    return jsonify({"message": "Lab Dip added successfully"}), 201

@app.route('/lab_dips/<int:dip_id>/status', methods=['PUT'])
def update_lab_dip_status(dip_id):
    data = request.json
    lab_dip = LabDip.query.get(dip_id)
    
    if not lab_dip:
        return jsonify({"error": "Lab Dip not found"}), 404

    lab_dip.status = data['status']
    if data['status'] == 'approved':
        lab_dip.approval_date = datetime.utcnow()

    db.session.commit()
    return jsonify({"message": "Lab Dip status updated successfully"}), 200

app.config["UPLOAD_FOLDER"] = "uploads"

@app.route("/upload_files", methods=["POST"])
def upload_files():
    buyer_name = request.form.get("buyerName")
    garment = request.form.get("garment")

    if not buyer_name or not garment:
        return jsonify({"message": "Buyer Name and Garment are required!"}), 400

    uploaded_files = {}
    for file_key in ["techpack", "bom", "specSheet"]:
        if file_key in request.files:
            file = request.files[file_key]
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], buyer_name + "_" + garment + "_" + file.filename )
            file.save(file_path)
            uploaded_files[file_key] = file_path

    # If a Techpack file was uploaded, create a new Style entry
    if "techpack" and "bom" in uploaded_files:
        new_style = Style(
            style_number="EBOLT-S-05",
            brand=buyer_name,
            sample_type="Fit Sample",
            garment=garment,
            color="Ivory",
            quantity="2",
            smv="42",
            order_received_date=datetime.today(),
            techpack_data = generateTechPackDataFromAi()
        )
        db.session.add(new_style)
        db.session.commit()

        # style_number_to_update = new_style.style_number
        # new_techpack_data = {
        #     "fabric": "Linen",
        #     "buttons": 5,
        #     "zippers": 1
        # }

        # db.session.query(Style).filter_by(style_number=style_number_to_update).update({
        #     "techpack_data": new_techpack_data
        # })
        # db.session.commit()


    return jsonify({"message": "Files uploaded successfully!", "files": uploaded_files})

@app.route("/upload_files_new", methods=["POST"])
def upload_files_new():
    buyer_name = request.form.get("buyerName")
    garment = request.form.get("garment")
    
    if not buyer_name or not garment:
        return jsonify({"message": "Buyer Name and Garment are required!"}), 400
    
    uploaded_files = {}
    fileNames = []
    for file_key in ["techpack", "bom", "specSheet"]:
        if file_key in request.files:
            file = request.files[file_key]
            if file.filename:  # Check if file is actually selected
                # Add file type indicator in the filename
                file_path = os.path.join(
                    app.config["UPLOAD_FOLDER"], 
                    f"{buyer_name}_{garment}_{file_key}_{file.filename}"
                )
                file.save(file_path)
                uploaded_files[file_key] = file_path
                fileNames.append(file.filename)
    
    # If at least Techpack and BOM files were uploaded, create a new Style entry
    if "techpack" in uploaded_files and "bom" in uploaded_files:
        # Generate style metadata and techpack data using AI
        ai_result = generateTechPackDataFromAi(buyer_name,garment,fileNames)
        style_metadata = ai_result.get("style_metadata", {})
        techpack_data = ai_result.get("techpack_data", {})
        
        # Create new style with the AI-generated data
        # Use AI-extracted values with fallbacks to form values or defaults
        new_style = Style(
            style_number=style_metadata.get("style_number") or "EBOLT-S-05",
            brand=style_metadata.get("brand") or buyer_name,
            sample_type=style_metadata.get("sample_type") or "Fit Sample",
            garment=style_metadata.get("garment") or garment,
            color=style_metadata.get("color") or techpack_data.get("shade") or "Ivory",
            quantity=style_metadata.get("quantity") or "2",
            smv=style_metadata.get("smv") or "42",
            order_received_date=datetime.today(),
            techpack_data=techpack_data
        )
        
        db.session.add(new_style)
        db.session.commit()
        sample_tracking_sample = SampleTrackerStyle.query.filter_by(style_number=style_metadata.get("style_number")).first()

        if not sample_tracking_sample:
            new_sample_style = SampleTrackerStyle(
                style_number=style_metadata.get("style_number"),
                brand=style_metadata.get("brand"),
                garment_type=style_metadata.get("garment"),
                start_date=datetime.now(timezone.utc) # Set start date on creation
            )
            db.session.add(new_sample_style)
            db.session.flush()  # Flush to get the new_style.id for linking samples

            # Add predefined samples to this new style
            for sample_type_name in PREDEFINED_SAMPLE_TYPES:
                sample = SampleTrackerSample(
                    style_id=new_sample_style.id,
                    type=sample_type_name,
                    completed=False
                    # start_date and end_date will be null initially
                )
                db.session.add(sample)

            db.session.commit()

        
        
        return jsonify({
            "message": "Files uploaded and style created successfully!",
            "files": uploaded_files,
            "style": {
                "id": new_style.id,
                "style_number": new_style.style_number,
                "brand": new_style.brand,
                "garment": new_style.garment,
                "color": new_style.color
            }
        })
    
    return jsonify({
        "message": "Files uploaded successfully! No style created (requires techpack and BOM).", 
        "files": uploaded_files
    })

def generateTechPackData():
         new_techpack_data = {
    "shade": "Ivory",
    "patternNo" : "290638",
    "season": "SS25",
    "mainBodyFabric": "Jersey Waffle 100% Cotton",
    "collarFabric": "",
    "mainLabel": "EA02A00319CH - WHITE/BLUE ",
    "threadShade": "BLACK, TKT120 ",
    "sweingThreads": "STIPBPLAHUB1019 ,STILASATHUB2424 ",
    "sweingThreadsDetails": "POLY POLY-2994-EPIC-Tex 24 - TKT 120 , POLY POLY-2994-EPIC-Tex 18 - TKT 180",
    "costSheet": {
        "fabricCost": [
        {
            "fabricType": "Shell Fabric",
            "description": "Jersey Waffle 100% Cotton "
        }
        ],
       "trimCost": [
      {
        "trim": "BUTTON ",
        "descirption": "20134935 - 4 Holes - Gritt ",
        "quantity" : "3"
      },
      {
        "trim": "Interlining",
        "descirption": "Fusible interlining 9510 ",
        "quantity" : "66"
      },
      {
        "trim": "pocket label",
        "descirption": "Brand label OUTSIDE FLAG LABEL ",
        "quantity" : "4.25"
      },
      {
        "trim": "SIZE ",
        "descirption": "Size label LEISURE HUGO BLUE ",
        "quantity" : "13.6"
      },
      {
        "trim": "Main label",
        "descirption": "Brand label 30X44 (59) MM ",
        "quantity" : "7.65"
      },
      {
        "trim": "Care label ",
        "descirption": "Care label BEDRUCKT 35mm Breite",
        "quantity" : "14.28"
      },
      {
        "trim": "COO LABEL",
        "descirption": "Made in India",
        "quantity" : "5"
      },
      {
        "trim": "Thread",
        "descirption": "POLY POLY-2994-EPIC-Tex 24 - TKT 120",
        "quantity" : "177"
      },
      {
        "trim": "Thread",
        "descirption": "POLY POLY-2994-EPIC-Tex 18 - TKT 180",
        "quantity" : "174"
      }
    ]
    },
    "bom": {
        "fabric": [
        {
            "code": "",
            "description": "Jersey Waffle 100% Cotton",
            "color": "Ivory",
            "size": "M",
            "quantity": "1"
        }
        ],
        "trims": [
        {
            "code": "STIINCOTGEN5877",
            "trim": "Interlining",
            "descirption": "Fusible interlining 9510 ",
            "color": "9510 COL.3216 BLACK",
            "size": "M",
            "quantity": "66.00"
        },
        {
            "code": "STILAWOVHUB4475",
            "trim": "pocket label",
            "descirption": "Brand label OUTSIDE FLAG LABEL",
            "color": "EA02A00319CH - WHITE/BLUE",
            "size": "M",
            "quantity": "4.25"
        },
        {
            "code": "SSTILAWOVHUB4649",
            "trim": "SIZE",
            "descirption": "Size label LEISURE HUGO BLUE",
            "color": "EA02A00319CH - WHITE/BLUE",
            "size": "M",
            "quantity": "13.60"
        },
        {
            "code": "STILAWOVHUB1267",
            "trim": "Main label",
            "descirption": "Brand label 30X44 (59) MM",
            "color": "EA02A00319CH - WHITE/BLUE",
            "size": "M",
            "quantity": "7.65"
        },
        {
            "code": "STISTPAPHUB8475",
            "trim": "Care label",
            "descirption": "Care label BEDRUCKT 35mm Breite",
            "color": "CL31 CL32",
            "size": "M",
            "quantity": "14.28"
        },
        {
            "code": "STIHTPAPHUB1228",
            "trim": "Coo Label",
            "descirption": "Made in India",
            "color": "CL31 CL32",
            "size": "M",
            "quantity": "5.00"
        },
        {
            "code": "STIPBPLAHUB1019",
            "trim": "Thread",
            "descirption": "POLY POLY-2994-EPIC-Tex 24 - TKT 120",
            "color": "2994-TEX-24-POLY POLY-SH-, BLACK, TKT120",
            "size": "M",
            "quantity": "177.00"
        },
        {
            "code": "STILASATHUB2424",
            "trim": "Thread",
            "descirption": "POLY POLY-2994-EPIC-Tex 18 - TKT 180",
            "color": "2994-TEX-24-POLY POLY-SH-, BLACK, TKT120",
            "size": "M",
            "quantity": "174.00"
        }
        ]
    }
        }
         return new_techpack_data


@app.route("/get_uploaded_files", methods=["GET"])
def get_uploaded_files():
    file_list = []
    upload_dir = app.config["UPLOAD_FOLDER"]
    
    for file_name in os.listdir(upload_dir):
        parts = file_name.split("_")
        if len(parts) >= 3:
            buyer_name, garment, file_type = parts[0], parts[1], parts[2]
            file_list.append({
                "buyerName": buyer_name,
                "garment": garment,
                "fileType": file_type,
                "filePath": f"/uploads/{file_name}"
            })
    
    return jsonify(file_list)


@app.route('/api/trims', methods=['GET'])
def get_trims():
    trims = Trim.query.all()
    return jsonify([{
        'id': trim.id,
        'name': trim.name,
        'image': trim.image,
        'composition': trim.composition,
        'structure': trim.structure,
        'shade': trim.shade,
        'brand': trim.brand,
        'code': trim.code
    } for trim in trims])

# Get a single trim by name
@app.route('/api/trims/<string:trim_name>', methods=['GET'])
def get_trim(trim_name):
    trim = Trim.query.filter_by(name=trim_name).first()
    if not trim:
        return jsonify({'error': 'Trim not found'}), 404
    trimVariants = TrimVariant.query.filter_by(trim_id = trim.id)
    variants = [
        {
            "id": variant.id,
            "image": variant.image,
            "composition": variant.composition,
            "structure": variant.structure,
            "shade": variant.shade,
            "brand": variant.brand,
            "code" : variant.code,
            "rate" : variant.rate
        }
        for variant in trimVariants
    ]


    return jsonify({
        'id': trim.id,
        'name': trim.name,
        'image': trim.image,
        'composition': trim.composition,
        'structure': trim.structure,
        'shade': trim.shade,
        'brand': trim.brand,
        'code': trim.code,
        "variants": variants
    })

# Add a new trim
@app.route('/api/trims', methods=['POST'])
def add_trim():
    data = request.json
    new_trim = Trim(
        name=data['name'],
        image=data.get('image', ''),
        composition=data.get('composition', ''),
        structure=data.get('structure', ''),
        shade=data.get('shade', ''),
        brand=data.get('brand', ''),
        code=data.get('code', '')
        )
    db.session.add(new_trim)
    db.session.commit()
    return jsonify({'message': 'Trim added successfully'}), 201

# Delete a trim
@app.route('/api/variants/<int:variant_id>', methods=['DELETE'])
def delete_variant(variant_id):
    variant = TrimVariant.query.get(variant_id)
    if not variant:
        return jsonify({'error': 'Variant not found'}), 404

    db.session.delete(variant)
    db.session.commit()
    return jsonify({'message': 'Variant deleted successfully'})


# Update a trim
@app.route('/api/trims/<string:trim_name>', methods=['PUT'])
def update_trim(trim_name):
    trim = Trim.query.filter_by(name=trim_name).first()
    if not trim:
        return jsonify({'error': 'Trim not found'}), 404

    data = request.json
    trim.image = data.get('image', trim.image)
    trim.composition = data.get('composition', trim.composition)
    trim.structure = data.get('structure', trim.structure)
    trim.shade = data.get('shade', trim.shade)
    trim.brand = data.get('brand', trim.brand)
    trim.code = data.get('code', trim.code)
    trim.rate = data.get('rate',trim.rate)

    db.session.commit()
    return jsonify({'message': 'Trim updated successfully'})

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


### ✅ POST API to Add a Trim Variant
@app.route('/api/trims/<trim_id>/variants', methods=['POST'])
def add_trim_variant(trim_id):
    trim = Trim.query.get(trim_id)
    if not trim:
        return jsonify({"error": "Trim not found"}), 404

    data = request.json
    image = data.get("image")
    composition = data.get("composition")
    structure = data.get("structure")
    shade = data.get("shade")
    brand = data.get("brand")
    code = data.get("code")
    rate = data.get("rate",'')


    if not image or not composition or not structure or not shade or not brand:
        return jsonify({"error": "All fields are required"}), 400

    new_variant = TrimVariant(
        trim_id=trim.id,
        image=image,
        composition=composition,
        structure=structure,
        shade=shade,
        brand=brand,
        code = code,
        rate = rate
    )

    db.session.add(new_variant)
    db.session.commit()

    return jsonify({"message": "Trim variant added successfully", "id": new_variant.id})

# ✅ Fetch notifications (GET)
@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    notifications = Notification.query.order_by(Notification.timestamp.desc()).all()
    return jsonify([n.to_dict() for n in notifications])

# ✅ Add a new notification (POST)
@app.route('/api/notifications', methods=['POST'])
def create_notification():
    data = request.json
    if 'message' not in data:
        return jsonify({"error": "Message is required"}), 400

    new_notification = Notification(message=data['message'])
    db.session.add(new_notification)
    db.session.commit()

    return jsonify(new_notification.to_dict()), 201

# ✅ Mark notification as read (PATCH)
@app.route('/api/notifications/<int:id>/read', methods=['PATCH'])
def mark_notification_as_read(id):
    notification = Notification.query.get(id)
    if not notification:
        return jsonify({"error": "Notification not found"}), 404

    notification.read_status = True
    db.session.commit()
    return jsonify(notification.to_dict())

# ✅ Delete a notification (DELETE)
@app.route('/api/notifications/<int:id>', methods=['DELETE'])
def delete_notification(id):
    notification = Notification.query.get(id)
    if not notification:
        return jsonify({"error": "Notification not found"}), 404

    db.session.delete(notification)
    db.session.commit()
    return jsonify({"message": "Notification deleted successfully"})

@app.route('/api/notifications/clear-all', methods=['DELETE'])
def delete_all_notification():
    notifications = Notification.query.filter_by(read_status=False).all()
    if not notifications:
        return jsonify({"error": "Notification not found"}), 404

    for notification in notifications:  # Iterate through the list
        db.session.delete(notification) 
    db.session.commit()
    return jsonify({"message": "Notifications deleted successfully"})



def query_gemini(prompt):
    """Send a query to Gemini and return the response."""
    model = genai.GenerativeModel("gemini-2.0-flash")
    response = model.generate_content(prompt)
    return response.text

def generate_ai_response(query, data):
    model = genai.GenerativeModel("gemini-2.0-flash")
    # Use the custom encoder for JSON serialization
    json_data = json.dumps(data, indent=2, cls=CustomJSONEncoder)
    prompt = f"User Query: {query}\n\nRelevant Data: {json_data}\n\nProvide a helpful response."
    response = model.generate_content(prompt)
    return response.text.strip()


def generateTechPackDataFromAi(buyer_name,garment,fileNames):
    """
    Process uploaded techpack and BOM files using Google's Gemini LLM
    to extract and structure the techpack data along with style details.
    Returns both style metadata and techpack technical data.
    """

    # Get the current request's files from upload folder
    upload_folder = app.config["UPLOAD_FOLDER"] # Make sure 'app' is defined in this scope
    
    # Find the most recently uploaded techpack and BOM files
    files = {}
    for filename in os.listdir(upload_folder):
        if "_techpack_" in filename.lower() and os.path.isfile(os.path.join(upload_folder, filename)) and any(name in filename for name in fileNames):
            files["techpack"] = os.path.join(upload_folder, filename)
        elif "_bom_" in filename.lower() and os.path.isfile(os.path.join(upload_folder, filename)) and any(name in filename for name in fileNames):
            files["bom"] = os.path.join(upload_folder, filename)
        elif "_specSheet_" in filename.lower() and os.path.isfile(os.path.join(upload_folder, filename)):
            files["specSheet"] = os.path.join(upload_folder, filename)
    trims_data = db.session.query(Trim.name, TrimVariant).join(TrimVariant, Trim.id == TrimVariant.trim_id).all()
    trim_variant_list = []

    for trim_name, variant in trims_data:
        trim_variant_list.append({
            "trim_name": trim_name,
            "id": variant.id,
            "composition": variant.composition,
            "structure": variant.structure,
            "shade": variant.shade,
            "brand": variant.brand,
            "code": variant.code,
            "rate": variant.rate
        })
    fabric_variant_list = []

    fabric_data = db.session.query(Fabric.name, FabricVariant).join(FabricVariant, Fabric.id == FabricVariant.fabric_id).all()
    for fabric_name, variant in fabric_data:
        fabric_variant_list.append({
            "fabric_name": fabric_name,
            "id": variant.id,
            "composition": variant.composition,
            "structure": variant.structure,
            "shade": variant.shade,
            "brand": variant.brand,
            "code": variant.code,
            "rate": variant.rate,
            "supplier" : variant.supplier
        })


    # Read the file contents based on file type
    file_contents = {}
    for file_type, file_path in files.items():
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            
            # Handle PDF files
            if file_extension == '.pdf':
                pdf_text = extract_text_from_pdf(file_path) # Make sure this function is defined
                file_contents[file_type] = pdf_text
            
            # Handle Excel files (xlsx, xls)
            elif file_extension in ['.xlsx', '.xls']:
                excel_text = extract_text_from_excel(file_path) # Make sure this function is defined
                file_contents[file_type] = excel_text
            
            # Handle text files (txt, csv, etc.)
            else:
                with open(file_path, 'r', encoding='utf-8') as file:
                    file_contents[file_type] = file.read()
                    
        except Exception as e:
            # app.logger.error(f"Error reading {file_type} file: {e}") # Make sure 'app.logger' is available
            print(f"Error reading {file_type} file: {e}") # Using print for standalone example
            file_contents[file_type] = f"Error reading file: {e}"
    
    # Create prompt for the LLM
    # MODIFIED PROMPT: Escaped curly braces for the JSON example
    prompt = '''
    Analyze the provided techpack, BOM (Bill of Materials),trims data, and extract the following information in a structured JSON format:
    
    1. Style information: style_number, brand, sample_type, garment type, color, quantity, smv (Standard Minute Value, if no smv present mark it as 42)
    2. Basic product information: shade, patternNo, season, mainBodyFabric, collarFabric, mainLabel, threadShade, sewingThreads, sewingThreadsDetails
    3. Cost sheet information with fabric costs and trim costs
    4. BOM (Bill of Materials) information including fabrics and trims
    5. After extracting BOM and Techpack data analyze the trims data and populate the 'costSheet' key in the format described below by matching all trims and fabrics code and taking rate from trims data. 
    6. For Quantity use the Consumption value instead of quantity as consumption is quantity per garment which is what we are capturing here
    Format the response exactly like this example:
    {{  
        "style_metadata": {{
            "style_number": "EBOLT-S-05",
            "brand": "BrandName",
            "sample_type": "Fit Sample",
            "garment": "Polo Shirt",
            "color": "Ivory",
            "quantity": "2",
            "smv": "42"
        }},
        "techpack_data": {{
            "shade": "Ivory",
            "patternNo": "290638",
            "season": "SS25",
            "mainBodyFabric": "Jersey Waffle 100% Cotton",
            "collarFabric": "",
            "mainLabel": "EA02A00319CH - WHITE/BLUE",
            "threadShade": "BLACK, TKT120",
            "sewingThreads": "STIPBPLAHUB1019, STILASATHUB2424",
            "sewingThreadsDetails": "POLY POLY-2994-EPIC-Tex 24 - TKT 120, POLY POLY-2994-EPIC-Tex 18 - TKT 180",
            "costSheet": {{
                "fabricCost": [
                    {{
                        "fabricType": "Shell Fabric",
                        "description": "Jersey Waffle 100% Cotton",
                        "quantity": "3",
                        "rate" : 242
                    }}
                ],
                "trimCost": [
                    {{
                        "trim": "BUTTON",
                        "description": "20134935 - 4 Holes - Gritt",
                        "quantity": "3",
                        "rate" : 242
                    }},
                    {{
                        "trim": "Interlining",
                        "description": "Fusible interlining 9510",
                        "quantity": "66",
                        "rate" : 242
                    }}
                    // other trim costs...
                ]
            }},
            "bom": {{
                "fabric": [
                    {{
                        "code": "STIINCOTGEN581111",
                        "description": "Jersey Waffle 100% Cotton",
                        "color": "Ivory",
                        "size": "M",
                        "quantity": "1",
                        "supplier" : "MANOHAR FILAMENT"
                    }}
                ],
                "trims": [
                    {{
                        "code": "STIINCOTGEN5877",
                        "trim": "Interlining",
                        "description": "Fusible interlining 9510",
                        "color": "9510 COL.3216 BLACK",
                        "size": "M",
                        "quantity": "66.00",
                        "supplier" : "MANOHAR FILAMENT"
                    }}
                    // other trims...
                ]
            }}
        }}
    }}
    
    If any information is not available in the provided documents, make an educated guess based on the available data or leave it as an empty string.
    
    Return ONLY the JSON with no explanations or additional text.
    
    Techpack data:
    {0}
    
    BOM data:
    {1}

    Trims data:
    {2}
    Fabric data:
    {3}
    If you are unable to fetch buyer name from provided data use this: {4}
    If you are unable to extract garment type from provided data user this : {5}
    ''' # You could add {2} here if you plan to include specSheet_data in the format call
    
    # Format the prompt with the file contents
    formatted_prompt = prompt.format(
        file_contents.get("techpack", "No techpack file found"),
        file_contents.get("bom", "No BOM file found"),
        trim_variant_list,
        fabric_variant_list,
        buyer_name,
        garment
        # If you want to include specSheet data, uncomment the line below
        # and add a {2} placeholder in the prompt string.
        # file_contents.get("specSheet", "No spec sheet found") 
    )
    
    try:
        # Generate response using Gemini
        model = genai.GenerativeModel("gemini-2.0-flash") # Ensure genai is imported and configured
        response = model.generate_content(formatted_prompt)
        response_text = response.text.strip().removeprefix("```json").removesuffix('```')
        app.logger.info(response_text)
        # For testing without API call, using the example JSON structure
        # response_text = """
        # {
        #     "style_metadata": {
        #         "style_number": "EBOLT-S-05",
        #         "brand": "BrandName",
        #         "sample_type": "Fit Sample",
        #         "garment": "Polo Shirt",
        #         "color": "Ivory",
        #         "quantity": "2",
        #         "smv": "42"
        #     },
        #     "techpack_data": {
        #         "shade": "Ivory",
        #         "patternNo": "290638",
        #         "season": "SS25",
        #         "mainBodyFabric": "Jersey Waffle 100% Cotton",
        #         "collarFabric": "",
        #         "mainLabel": "EA02A00319CH - WHITE/BLUE",
        #         "threadShade": "BLACK, TKT120",
        #         "sewingThreads": "STIPBPLAHUB1019, STILASATHUB2424",
        #         "sewingThreadsDetails": "POLY POLY-2994-EPIC-Tex 24 - TKT 120, POLY POLY-2994-EPIC-Tex 18 - TKT 180",
        #         "costSheet": {
        #             "fabricCost": [
        #                 {
        #                     "fabricType": "Shell Fabric",
        #                     "description": "Jersey Waffle 100% Cotton"
        #                 }
        #             ],
        #             "trimCost": [
        #                 {
        #                     "trim": "BUTTON",
        #                     "descirption": "20134935 - 4 Holes - Gritt",
        #                     "quantity": "3"
        #                 },
        #                 {
        #                     "trim": "Interlining",
        #                     "descirption": "Fusible interlining 9510",
        #                     "quantity": "66"
        #                 }
        #             ]
        #         },
        #         "bom": {
        #             "fabric": [
        #                 {
        #                     "code": "",
        #                     "description": "Jersey Waffle 100% Cotton",
        #                     "color": "Ivory",
        #                     "size": "M",
        #                     "quantity": "1"
        #                 }
        #             ],
        #             "trims": [
        #                 {
        #                     "code": "STIINCOTGEN5877",
        #                     "trim": "Interlining",
        #                     "descirption": "Fusible interlining 9510",
        #                     "color": "9510 COL.3216 BLACK",
        #                     "size": "M",
        #                     "quantity": "66.00"
        #                 }
        #             ]
        #         }
        #     }
        # }
        # """
        # response_text = response_text.strip()

        # Ensure response is valid JSON
        try:
            full_data = json.loads(response_text) # Ensure json is imported
            
            # Extract style metadata and techpack data
            style_metadata = full_data.get("style_metadata", {})
            techpack_data = full_data.get("techpack_data", {})
            
            # Fix any common formatting issues in techpack data
            if "sewingThreads" in techpack_data:
                # Typo was sweingThreads, should be sewingThreads
                # techpack_data["sweingThreads"] = techpack_data.pop("sewingThreads") 
                pass # Keeping as sewingThreads if model returns it correctly
            if "sewingThreadsDetails" in techpack_data:
                # Typo was sweingThreadsDetails, should be sewingThreadsDetails
                # techpack_data["sweingThreadsDetails"] = techpack_data.pop("sewingThreadsDetails")
                pass # Keeping as sewingThreadsDetails if model returns it correctly
                
            # Fix any 'description' that was spelled as 'descirption'
            if "costSheet" in techpack_data and "trimCost" in techpack_data["costSheet"]:
                for item in techpack_data["costSheet"]["trimCost"]:
                    if "descirption" in item:
                        item["description"] = item.pop("descirption")
            
            if "bom" in techpack_data and "trims" in techpack_data["bom"]:
                for item in techpack_data["bom"]["trims"]:
                    if "descirption" in item:
                        item["description"] = item.pop("descirption")
            
            # Combine metadata and technical data
            result = {
                "style_metadata": style_metadata,
                "techpack_data": techpack_data
            }
            
            return result
            
        except json.JSONDecodeError:
            # app.logger.error("AI response was not valid JSON")
            print("AI response was not valid JSON")
            # Return a minimal valid structure if parsing fails
            return {
                "style_metadata": {
                    "style_number": "", "brand": "", "sample_type": "Fit Sample", 
                    "garment": "", "color": "", "quantity": "1", "smv": ""
                },
                "techpack_data": {
                    "shade": "", "patternNo": "", "season": "", "mainBodyFabric": "",
                    "collarFabric": "", "mainLabel": "", "threadShade": "",
                    "sewingThreads": "", "sewingThreadsDetails": "", # Corrected typo
                    "costSheet": {"fabricCost": [], "trimCost": []},
                    "bom": {"fabric": [], "trims": []}
                }
            }
    
    except Exception as e:
        # app.logger.error(f"Error generating techpack data with AI: {e}")
        print(f"Error generating techpack data with AI: {e}")
        # Return a minimal valid structure if AI processing fails
        return {
            "style_metadata": {
                "style_number": "", "brand": "", "sample_type": "Fit Sample", 
                "garment": "", "color": "", "quantity": "1", "smv": ""
            },
            "techpack_data": {
                "shade": "", "patternNo": "", "season": "", "mainBodyFabric": "",
                "collarFabric": "", "mainLabel": "", "threadShade": "",
                "sewingThreads": "", "sewingThreadsDetails": "", # Corrected typo
                "costSheet": {"fabricCost": [], "trimCost": []},
                "bom": {"fabric": [], "trims": []}
            }
        }

def extract_text_from_pdf(pdf_path):
    """
    Extract text content from a PDF file
    """
    text = ""
    try:
        # Open the PDF file in binary mode
        with open(pdf_path, 'rb') as file:
            # Create a PDF reader object
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Extract text from each page
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n\n"
    except Exception as e:
        text = f"Error extracting text from PDF: {str(e)}"
    
    return text

def extract_text_from_excel(excel_path):
    """
    Extract text content from an Excel file (xlsx or xls), attempting to detect a meaningful table
    by matching expected column headers. Falls back to default logic if no match is found.
    """
    text = ""
    expected_headers = [
        "Name", "Order Qty", "SIZE", "Description", "supplier", "COLOUR",
        "SAP Item Code", "Consumption", "Total Required Qty", "UOM",
        "PR Trims Released date-26.11.2024 &  Fabric-28.11.2024", "SAP PR QTY",
        "PO QTY", "PO NUMBERS", "PO DATE", "PRICE", "VALUE", "INHOUSE DATE", "Remarks"
    ]

    try:
        # Try using pandas first
        try:
            xls = pd.ExcelFile(excel_path)
            sheet_names = xls.sheet_names

            for sheet_name in sheet_names:
                df_raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

                # Try to find a header row with partial match
                header_row_index = None
                for idx, row in df_raw.iterrows():
                    row_str = row.astype(str).str.lower().tolist()
                    match_count = sum(1 for header in expected_headers if header.lower() in row_str)
                    if match_count >= 5:
                        header_row_index = idx
                        break

                text += f"Sheet: {sheet_name}\n"
                
                if header_row_index is not None:
                    df_table = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row_index)
                    df_table_cleaned = df_table.dropna(how='all')
                    text += df_table_cleaned.to_string(index=False) + "\n\n"
                else:
                    # Fall back to default if no header match found
                    df_fallback = pd.read_excel(excel_path, sheet_name=sheet_name)
                    text += df_fallback.to_string(index=False) + "\n\n"

        # If pandas fails, use openpyxl
        except:
            workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"

                for row in sheet.rows:
                    row_text = " | ".join([str(cell.value) if cell.value is not None else "" for cell in row])
                    text += row_text + "\n"
                text += "\n"

    except Exception as e:
        text = f"Error extracting text from Excel: {str(e)}"

    return text

@app.route('/api/activity/update', methods=['POST'])
def update_activity_from_progress():
    data = request.json
    style = data.get('style')
    process = data.get('process')
    is_checked = data.get('isChecked')
    next_process = data.get('nextProcess')
    
    current_time = datetime.now().strftime('%Y-%m-%d')
    
    # Find the activity record for the current process
    activity = Activity.query.filter_by(style=style, process=process).first()
    
    if not activity:
        return jsonify({"error": f"No activity found for style {style} and process {process}"}), 404
    
    if is_checked:
        # Update the actual_end time for the current process
        activity.actual_end = current_time
        
        # Calculate actual duration if actual_start exists
        if activity.actual_start:
            actual_start_time = datetime.strptime(activity.actual_start, '%Y-%m-%d')
            actual_end_time = datetime.strptime(current_time, '%Y-%m-%d')
            activity.actual_duration = (actual_end_time - actual_start_time).total_seconds() / 3600  # Convert to hours
            
            # Calculate delay if planned_end exists
            if activity.planned_end:
                planned_end_time = datetime.strptime(activity.planned_end, '%Y-%m-%d')
                activity.delay = (actual_end_time - planned_end_time).total_seconds() / 3600  # Convert to hours
        
        # If there's a next process, update its actual_start time
        if next_process:
            next_activity = Activity.query.filter_by(style=style, process=next_process).first()
            if next_activity:
                next_activity.actual_start = current_time
                db.session.add(next_activity)
    else:
        # If checkbox is unchecked, clear the actual_end time
        activity.actual_end = None
        activity.actual_duration = None
        activity.delay = None
        
        # Also clear the actual_start time of the next process
        if next_process:
            next_activity = Activity.query.filter_by(style=style, process=next_process).first()
            if next_activity:
                next_activity.actual_start = None
                db.session.add(next_activity)
    
    db.session.add(activity)
    db.session.commit()
    
    return jsonify({"message": "Activity updated successfully"}), 200

@app.route('/api/activity/<style>', methods=['GET'])
def get_activities_from_progress(style):
    activities = Activity.query.filter_by(style=style).all()
    
    result = []
    for activity in activities:
        result.append({
            'id': activity.id,
            'style': activity.style,
            'process': activity.process,
            'duration': activity.duration,
            'planned_start': activity.planned_start,
            'planned_end': activity.planned_end,
            'actual_start': activity.actual_start,
            'actual_end': activity.actual_end,
            'actual_duration': activity.actual_duration,
            'delay': activity.delay,
            'responsibility': activity.responsibility
        })
    
    return jsonify(result), 200

@app.route('/api/activity/create', methods=['POST'])
def create_activities_from_progress():
    data = request.json
    style = data.get('style')
    processes = data.get('processes')
    
    if not style or not processes:
        return jsonify({"error": "Style and processes are required"}), 400
    
    # Default values for new activities
    current_time = datetime.now().strftime('%Y-%m-%d')
    
    created_activities = []
    
    for i, process in enumerate(processes):
        # Calculate planned times based on sequence
        planned_start = current_time  # Default start time
        
        # Simple logic to stagger planned times - adjust as needed
        if i > 0:
            previous_planned_end = datetime.strptime(planned_start, '%Y-%m-%d')
            planned_start = (previous_planned_end + datetime.timedelta(hours=2)).strftime('%Y-%m-%d')
        
        planned_end = (datetime.strptime(planned_start, '%Y-%m-%d') + 
                      datetime.timedelta(hours=4)).strftime('%Y-%m-%d')  # Default 4 hours per process
        
        # Create the activity record
        activity = Activity(
            style=style,
            process=process,
            duration=4.0,  # Default 4 hours duration
            planned_start=planned_start,
            planned_end=planned_end,
            responsibility="Production Team"  # Default responsibility
        )
        
        db.session.add(activity)
        
        created_activities.append({
            'style': style,
            'process': process,
            'planned_start': planned_start,
            'planned_end': planned_end
        })
    
    db.session.commit()
    
    return jsonify({
        "message": f"Created {len(created_activities)} activity records for style {style}",
        "activities": created_activities
    }), 201


@app.route('/chatbot/query', methods=['POST'])
def chatbot_query():
    data = request.json
    user_query = data.get("query", "").lower()

    response = "Sorry, I couldn't understand your query."

    # 1️⃣ Order Status
    if "status of style" in user_query or "order status" in user_query:
        style_no = user_query.split()[-1]  # Extract style number
        style_info  = Style.query.filter_by(style_number = style_no).first()
        task_info = Task.query.filter_by(style_number = style_no).first()
        activities_info = Activity.query.filter_by(style=style_no)
        response = get_response_string_from_style_info(style_info, task_info, activities_info)

    # 3️⃣ TNA Activity
    elif "activity status" in user_query:
        activities = Activity.query.filter(Activity.status != 'Completed').all()
        response = [{"Process": a.process, "Status": a.status} for a in activities]

    # 5️⃣ If no predefined query is matched, use Gemini for AI-generated responses
    elif "expected delivery date" in user_query:
        style_no = user_query.split()[-1]  # Extract style number
        style_info  = Style.query.filter_by(style_number = style_no).first()
        response = f"Expected Delivery Date for Style {style_info.style_number} is {style_info.order_delivery_date}"

    # elif "bom" or "bill of material" or "cost sheet" in user_query:
    #     style_no = user_query.split()[-1]  # Extract style number
    #     style_info  = Style.query.filter_by(style_number = style_no).first()
    #     techpack_data = style_info.techpack_data
    #     response = generate_ai_response(user_query, techpack_data)

    else:
        style_data = get_style_data_for_gemini(user_query)
        if style_data is not None:
            response =  generate_ai_response(f"User asked: {user_query}",style_data)
        else:
            response = "Sorry, I couldn't understand your query."

    return jsonify({"response": response})

def get_style_data_for_gemini(user_query):
    style_no = user_query.split()[-1]  # Extract style number
    style_info = Style.query.filter_by(style_number=style_no).first()
    
    # Convert style data to dictionary with date handling
    style_data = {}
    if style_info:
        style_data = {
            "id": style_info.id,
            "style": style_info.style_number,
            "buyer": style_info.brand,
            "garment": style_info.garment,
            "date": style_info.order_received_date.strftime('%d/%m/%Y') if style_info.order_received_date else None,
            "approvalStatus": style_info.approval_status,
            "order_delivery_date": style_info.order_delivery_date.strftime('%d/%m/%Y') if style_info.order_delivery_date else None,
            "techpack_data": style_info.techpack_data
        }
    else:
        return None
    # Convert activity data to list of dictionaries
    activity_data = Activity.query.filter_by(style=style_no).all()
    activity_list = [{
        "id": act.id,
        "style": act.style,
        "process": act.process,
        "duration": act.duration,
        "planned_start": act.planned_start,
        "planned_end": act.planned_end,
        "actual_start": act.actual_start,
        "actual_end": act.actual_end,
        "actual_duration": act.actual_duration,
        "delay": act.delay,
        "responsibility": act.responsibility
    } for act in activity_data]
    
    # Convert task data to list of dictionaries
    task_data = Task.query.filter_by(style_number=style_no).all()
    task_list = [{
        "id": task.id,
        "style_number": task.style_number,
        "brand": task.brand,
        "sample_type": task.sample_type,
        "garment": task.garment,
        "status": task.status,
        "progress": task.progress,
        "priority": task.priority,
        "timestamp": task.timestamp.strftime("%Y-%m-%d %H:%M:%S") if task.timestamp else None,
        "comment": task.comment,
        "problem_reported": task.problem_reported
    } for task in task_data]
    
    return {
        "style_data": style_data,
        "activity_data": activity_list,
        "task_data": task_list
    }

def get_response_string_from_style_info(style_info, task_info, activity_info : Activity):
    current_activity = None 
    for activity in activity_info:
        if activity.actual_end is None:
            current_activity = activity
            break

    return f"Style Number : {style_info.style_number} \n Brand: {style_info.brand} \n Task Status: {task_info.status} \n Activity Status: {current_activity.process}"

@app.route("/analyze-image", methods=["POST"])
def analyze_image():
    image_file = request.files.get("image")
    if not image_file:
        return jsonify({"error": "No image uploaded"}), 400

    try:
        # Read and process the image
        image_bytes = image_file.read()
        
        # Create PIL Image object
        pil_image = Image.open(io.BytesIO(image_bytes))
        
        # Convert to RGB if necessary (handles RGBA, etc.)
        if pil_image.mode != 'RGB':
            pil_image = pil_image.convert('RGB')

        model = genai.GenerativeModel("gemini-1.5-flash")

        prompt = """
You are a fashion image analysis assistant.

Analyze the garment in this image and provide your assessment in the following structured JSON format:

{
  "garmentType": "",
  "pattern": "",
  "color": "",
  "fit": "",
  "style": "",
  "collarType": "",
  "gender": "",
  "patternDistribution": {
    "Stripes": 0,
    "Checks": 0,
    "Solid": 0,
    "Floral": 0
  },
  "colorDistribution": {
    "Blue": 0,
    "Red": 0,
    "White": 0,
    "Black": 0
  }
}

Please analyze:
- The type of garment (shirt, dress, pants, etc.)
- The pattern (if any)
- Primary colors
- The fit and style
- Collar type (if applicable)
- Gender target (male/female/unisex)
- Pattern distribution as percentages (0-100)
- Color distribution as percentages (0-100)

Respond only with the JSON structure filled with your analysis.
        """

        # Generate content with both text prompt and image
        response = model.generate_content([prompt, pil_image])
        
        # Extract and return the JSON response
        return jsonify(extract_json(response.text))

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
def extract_json(text):
    import json, re
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        return json.loads(match.group())
    raise ValueError("No JSON found in response")

# Role-based access control decorators
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user or not user.is_admin():
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated_function

def role_required(role_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            current_user_id = get_jwt_identity()
            user = User.query.get(current_user_id)
            if not user or not user.has_role(role_name):
                return jsonify({"error": f"{role_name} access required"}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Audit logging function
def log_audit(user_id, action, table_name, record_id, old_values=None, new_values=None, ip_address=None):
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        table_name=table_name,
        record_id=record_id,
        old_values=old_values,
        new_values=new_values,
        ip_address=ip_address
    )
    db.session.add(audit_log)
    db.session.commit()

# Add endpoint to get audit logs (admin only)
@app.route('/api/audit-logs', methods=['GET'])
@jwt_required()
@admin_required
def get_audit_logs():
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).all()
    return jsonify([{
        'id': log.id,
        'user_id': log.user_id,
        'action': log.action,
        'table_name': log.table_name,
        'record_id': log.record_id,
        'old_values': log.old_values,
        'new_values': log.new_values,
        'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        'ip_address': log.ip_address
    } for log in logs])

# Fabric routes
@app.route('/api/fabrics', methods=['GET'])
def get_fabrics():
    """Get all fabrics"""
    fabrics = Fabric.query.all()
    return jsonify([{
        'id': fabric.id,
        'name': fabric.name,
        'image': fabric.image
    } for fabric in fabrics])

@app.route('/api/fabrics', methods=['POST'])
def create_fabric():
    """Create a new fabric"""
    data = request.get_json()
    
    # Check if fabric already exists
    existing_fabric = Fabric.query.filter_by(name=data['name']).first()
    if existing_fabric:
        return jsonify({'error': 'Fabric already exists'}), 400
    
    fabric = Fabric(
        name=data['name'],
        image=data.get('image', '')
    )
    
    db.session.add(fabric)
    db.session.commit()
    
    return jsonify({
        'id': fabric.id,
        'name': fabric.name,
        'image': fabric.image
    }), 201

@app.route('/api/fabrics/<fabric_name>', methods=['GET'])
def get_fabric_detail(fabric_name):
    """Get fabric details with variants"""
    fabric = Fabric.query.filter_by(name=fabric_name).first()
    if not fabric:
        return jsonify({'error': 'Fabric not found'}), 404
    
    variants = FabricVariant.query.filter_by(fabric_id=fabric.id).all()
    
    return jsonify({
        'id': fabric.id,
        'name': fabric.name,
        'image': fabric.image,
        'variants': [{
            'id': variant.id,
            'image': variant.image,
            'composition': variant.composition,
            'structure': variant.structure,
            'shade': variant.shade,
            'brand': variant.brand,
            'code': variant.code,
            'rate': variant.rate,
            'supplier' : variant.supplier
        } for variant in variants]
    })

@app.route('/api/fabrics/<fabric_name>', methods=['DELETE'])
def delete_fabric(fabric_name):
    """Delete a fabric and all its variants"""
    fabric = Fabric.query.filter_by(name=fabric_name).first()
    if not fabric:
        return jsonify({'error': 'Fabric not found'}), 404
    
    db.session.delete(fabric)
    db.session.commit()
    
    return jsonify({'message': 'Fabric deleted successfully'})

@app.route('/api/fabrics/<int:fabric_id>/variants', methods=['POST'])
def add_fabric_variant(fabric_id):
    """Add a variant to a fabric"""
    fabric = Fabric.query.get(fabric_id)
    if not fabric:
        return jsonify({'error': 'Fabric not found'}), 404
    
    data = request.get_json()
    
    variant = FabricVariant(
        fabric_id=fabric_id,
        image=data.get('image', ''),
        composition=data.get('composition', ''),
        structure=data.get('structure', ''),
        shade=data.get('shade', ''),
        brand=data.get('brand', ''),
        code=data.get('code', ''),
        rate=data.get('rate', ''),
        supplier = data.get('supplier', '')
    )
    
    db.session.add(variant)
    db.session.commit()
    
    return jsonify({
        'id': variant.id,
        'image': variant.image,
        'composition': variant.composition,
        'structure': variant.structure,
        'shade': variant.shade,
        'brand': variant.brand,
        'code': variant.code,
        'rate': variant.rate
    }), 201

@app.route('/api/fabric-variants/<int:variant_id>', methods=['DELETE'])
def delete_fabric_variant(variant_id):
    """Delete a fabric variant"""
    variant = FabricVariant.query.get(variant_id)
    if not variant:
        return jsonify({'error': 'Variant not found'}), 404
    
    db.session.delete(variant)
    db.session.commit()
    
    return jsonify({'message': 'Variant deleted successfully'})


# Initialize roles if they don't exist
def initialize_roles():
    roles = ['admin', 'manager', 'user']
    for role_name in roles:
        if not Role.query.filter_by(name=role_name).first():
            role = Role(name=role_name, description=f'{role_name.capitalize()} role')
            db.session.add(role)
    db.session.commit()

# Call initialize_roles when the application starts
with app.app_context():
    db.create_all()
    initialize_roles()

if __name__ == '__main__':
    if not os.path.exists(app.config["UPLOAD_FOLDER"]):
        os.makedirs(app.config["UPLOAD_FOLDER"])
    app.run(debug=True)
